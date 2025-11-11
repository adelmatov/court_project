import psycopg2
import psycopg2.extras
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from datetime import datetime, date
import hashlib
import os
import re
import logging
import sys
import codecs

# ==================== НАСТРОЙКА ЛОГИРОВАНИЯ ====================
# Установить UTF-8 для stdout/stderr в Windows
if sys.platform == "win32":
    # Обернуть stdout/stderr в UTF-8 writer
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'replace')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'replace')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ==================== НАСТРОЙКИ ====================
DB_QAMQOR = {
    'host': 'localhost',
    'database': 'qamqor',
    'user': 'postgres',
    'password': 'admin',
    'port': 5432
}

DB_COMPANIES = {
    'host': 'localhost',
    'database': 'companies',
    'user': 'postgres',
    'password': 'admin',
    'port': 5432
}

OUTPUT_FILE = r'C:\Users\adelmatov001\OneDrive - PwC\SharePoint List Update\qamqor_sharepoint_sync.xlsx'

# Все колонки (добавлена колонка Статус)
ALL_COLUMNS = [
    'Номер', 'Статус', 'Тип проверки', 'Дата регистрации', 'Дата акта', 
    'Дата начала', 'Дата окончания', 'Дата приостановки', 'Дата возобновления',
    'Начало продления', 'Конец продления', 'ОГД', 'КПССУ', 'Вид проверки',
    'БИН', 'Название субъекта', 'Адрес субъекта', 'Статус проверки', 'Тема аудита',
    'Тема аудита 1', 'Руководитель', 'КРП', 'КФК', 'КСЭ', 'ОКЭД',
    'Плательщик НДС', 'История налогов', 'Сумма налогов', 'История НДС', 'Сумма НДС'
]

# ==================== ФУНКЦИИ СОКРАЩЕНИЯ ====================

def shorten_revenue_name(name):
    """Сократить название органа доходов"""
    if not name or pd.isna(name):
        return ''
    
    name_lower = name.lower()
    
    if 'комитет государственных доходов' in name_lower and 'департамент' not in name_lower:
        return "КГД"
    
    if 'департамент государственных доходов' in name_lower:
        match = re.search(
            r'по\s+([А-ЯЁӘІҢҒҮҰҚӨҺа-яёәіңғүұқөһ\s\-]+?(?:ской|кой|нской|скай)\s+област[иь])',
            name, re.IGNORECASE
        )
        if match:
            region = match.group(1).strip()
            region = re.sub(r'скай\s+област', 'ской област', region, flags=re.IGNORECASE)
            
            region_abbr = {
                'Северо-Казахстанской области': 'СКО',
                'Восточно-Казахстанской области': 'ВКО',
                'Западно-Казахстанской области': 'ЗКО',
                'Южно-Казахстанской области': 'ЮКО',
            }
            
            for full_name, abbr in region_abbr.items():
                if region.lower() == full_name.lower():
                    return f"ДГД по {abbr}"
            
            return f"ДГД по {region}"
        
        match = re.search(
            r'по области\s+([А-ЯЁӘІҢҒҮҰҚӨҺа-яёәіңғүұқөһA-Za-z\s\-]+?)(?:\s+Комитета|\s*["\»])', 
            name, re.IGNORECASE
        )
        if match:
            region = match.group(1).strip()
            return f"ДГД по области {region}"
        
        if 'астан' in name_lower:
            return "ДГД по г. Астана"
        elif 'алматы' in name_lower:
            return "ДГД по г. Алматы"
        elif 'шымкент' in name_lower:
            return "ДГД по г. Шымкент"
    
    return name

def shorten_company_form(name):
    """Сократить организационно-правовую форму компании"""
    if not name or pd.isna(name):
        return ''
    
    replacements = {
        'Товарищество с ограниченной ответственностью': 'ТОО',
        'товарищество с ограниченной ответственностью': 'ТОО',
        'Товарищества с ограниченной ответственностью': 'ТОО',
        'Дочернее Товарищество с ограниченной ответственностью': 'Дочернее ТОО',
        'Дочернее товарищество с ограниченной ответственностью': 'Дочернее ТОО',
        'Акционерное общество': 'АО',
        'Закрытого акционерного общества': 'ЗАО',
        'Филиал Товарищества с ограниченной ответственностью': 'Филиал ТОО',
        'Филиал товарищества с ограниченной ответственностью': 'Филиал ТОО',
        'Филиал компании': 'Филиал',
        'Филиал Компании': 'Филиал',
        'Представительство Товарищества с ограниченной ответственностью': 'Представительство ТОО',
        'Представительство Компании': 'Представительство',
        'Представительство': 'Представительство',
        'Республиканское государственное учреждение': 'РГУ',
        'Государственное учреждение': 'ГУ',
        'Коммунальное государственное учреждение': 'КГУ',
        'Частная компания': 'ЧК',
        'Производственный кооператив': 'ПК',
        'производственный кооператив': 'ПК',
    }
    
    result = name
    for full_form, short_form in sorted(replacements.items(), key=lambda x: len(x[0]), reverse=True):
        if full_form in result:
            result = result.replace(full_form, short_form)
            break
    
    return result.strip()

def shorten_kpssu_name(name):
    """Сократить название КПССУ"""
    if not name or pd.isna(name):
        return ''
    
    long_patterns = [
        'Государственное учреждение "Департамент Комитета по правовой статистике и специальным учетам Генеральной прокуратуры Республики Казахстан',
        'Государственное учреждение «Департамент Комитета по правовой статистике и специальным учетам Генеральной прокуратуры Республики Казахстан',
        'Департамент Комитета по правовой статистике и специальным учетам Генеральной прокуратуры Республики Казахстан',
    ]
    
    result = name
    for pattern in long_patterns:
        if pattern in result:
            result = result.replace(pattern, 'КПССУ')
            break
    
    result = result.replace('""', '').replace('"', '').replace('«', '').replace('»', '')
    result = ' '.join(result.split())
    
    if result.startswith('по ') and 'КПССУ' not in result:
        result = 'КПССУ ' + result
    
    region_abbr = {
        'по Северо-Казахстанской области': 'по СКО',
        'по Восточно-Казахстанской области': 'по ВКО',
        'по Западно-Казахстанской области': 'по ЗКО',
        'по Южно-Казахстанской области': 'по ЮКО',
    }
    
    for full_name, abbr in region_abbr.items():
        if full_name in result:
            result = result.replace(full_name, abbr)
            break
    
    return result.strip()

# ==================== ФОРМАТИРОВАНИЕ ЧИСЕЛ ====================

def format_number_with_spaces(value):
    """Форматировать число с пробелами между разрядами"""
    if value is None or pd.isna(value):
        return ''
    
    formatted = f"{float(value):,.2f}"
    return formatted.replace(',', ' ')

# ==================== НОРМАЛИЗАЦИЯ ====================

def normalize_bin(bin_value):
    """Нормализовать БИН (дополнить нулями до 12 символов)"""
    if bin_value is None or pd.isna(bin_value):
        return ''
    
    bin_str = str(bin_value).strip()
    
    if bin_str.startswith("'"):
        bin_str = bin_str[1:]
    
    if bin_str.lower() in ['nan', 'none', 'nat', '']:
        return ''
    
    if '.' in bin_str:
        bin_str = bin_str.split('.')[0]
    
    return bin_str.zfill(12)

def normalize_value(value, is_bin=False):
    """Нормализовать значение для хеша"""
    if is_bin:
        return normalize_bin(value)
    
    if value is None or pd.isna(value):
        return ''
    
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in ['nan', 'none', 'nat', '']:
            return ''
        return stripped
    
    if isinstance(value, (date, datetime)):
        return value.strftime('%Y-%m-%d')
    
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    
    return str(value).strip()

def calculate_row_hash(row_dict):
    """Вычислить хеш строки"""
    exclude_fields = ['Номер', 'hash', 'Статус', 'Сумма налогов', 'Сумма НДС']
    
    values = []
    for key in sorted(row_dict.keys()):
        if key not in exclude_fields:
            is_bin = (key == 'БИН')
            values.append(f"{key}={normalize_value(row_dict[key], is_bin=is_bin)}")
    
    hash_string = '|'.join(values)
    return hashlib.md5(hash_string.encode('utf-8')).hexdigest()

# ==================== РАБОТА С БД ====================

def get_companies_data_batch(bins, conn):
    """Получить данные компаний пакетно"""
    if len(bins) == 0:
        return {}
    
    bins_list = list(bins)
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute("""
        SELECT 
            c.bin,
            c.ceo_name,
            rk.name as krp_name,
            rf.name as kfc_name,
            rks.name as kse_name,
            ro.name as oked_name,
            c.is_nds
        FROM companies c
        LEFT JOIN ref_krp rk ON c.krp_id = rk.id
        LEFT JOIN ref_kfc rf ON c.kfc_id = rf.id
        LEFT JOIN ref_kse rks ON c.kse_id = rks.id
        LEFT JOIN ref_oked ro ON c.oked_id = ro.id
        WHERE c.bin = ANY(%s)
    """, (bins_list,))
    
    companies = {row['bin']: dict(row) for row in cursor.fetchall()}
    
    cursor.execute("""
        SELECT bin, year, total_taxes
        FROM company_taxes
        WHERE bin = ANY(%s)
        ORDER BY bin, year DESC
    """, (bins_list,))
    
    tax_history = {}
    tax_totals = {}
    for row in cursor.fetchall():
        if row['bin'] not in tax_history:
            tax_history[row['bin']] = []
            tax_totals[row['bin']] = 0.0
        if row['year'] and row['total_taxes']:
            amount = float(row['total_taxes'])
            formatted_amount = format_number_with_spaces(amount)
            tax_history[row['bin']].append(f"{row['year']}: {formatted_amount} ₸")
            tax_totals[row['bin']] += amount
    
    cursor.execute("""
        SELECT bin, year, nds_amount
        FROM company_nds
        WHERE bin = ANY(%s)
        ORDER BY bin, year DESC
    """, (bins_list,))
    
    nds_history = {}
    nds_totals = {}
    for row in cursor.fetchall():
        if row['bin'] not in nds_history:
            nds_history[row['bin']] = []
            nds_totals[row['bin']] = 0.0
        if row['year'] and row['nds_amount']:
            amount = float(row['nds_amount'])
            formatted_amount = format_number_with_spaces(amount)
            nds_history[row['bin']].append(f"{row['year']}: {formatted_amount} ₸")
            nds_totals[row['bin']] += amount
    
    cursor.close()
    
    result = {}
    for bin_num in bins_list:
        if bin_num in companies:
            comp = companies[bin_num]
            result[bin_num] = {
                'ceo_name': normalize_value(comp['ceo_name']),
                'krp_name': normalize_value(comp['krp_name']),
                'kfc_name': normalize_value(comp['kfc_name']),
                'kse_name': normalize_value(comp['kse_name']),
                'oked_name': normalize_value(comp['oked_name']),
                'is_nds': 'Да' if comp['is_nds'] else 'Нет',
                'tax_history': '\n'.join(tax_history.get(bin_num, [])),
                'tax_total': tax_totals.get(bin_num, None),
                'nds_history': '\n'.join(nds_history.get(bin_num, [])),
                'nds_total': nds_totals.get(bin_num, None),
            }
        else:
            result[bin_num] = {
                'ceo_name': '',
                'krp_name': '',
                'kfc_name': '',
                'kse_name': '',
                'oked_name': '',
                'is_nds': '',
                'tax_history': '',
                'tax_total': None,
                'nds_history': '',
                'nds_total': None,
            }
    
    return result

def reset_all_statuses_to_current(file_path):
    """ШАГ 1: Сбросить все статусы на 'current' перед началом работы"""
    if not os.path.exists(file_path):
        logger.info("Файл не существует, пропускаем сброс статусов")
        return
    
    logger.info("Сброс всех статусов на 'current'...")
    
    wb = load_workbook(file_path)
    ws = wb['QamqorData']
    
    status_col_idx = 2
    
    updated_count = 0
    for row_idx in range(2, ws.max_row + 1):
        current_status = ws.cell(row=row_idx, column=status_col_idx).value
        if current_status != 'current':
            ws.cell(row=row_idx, column=status_col_idx).value = 'current'
            updated_count += 1
    
    wb.save(file_path)
    wb.close()
    
    logger.info(f"Сброшено статусов: {updated_count}")

def load_existing_data(file_path):
    """Загрузить существующие данные из Excel И вычислить hash для каждой строки"""
    if not os.path.exists(file_path):
        return pd.DataFrame()
    
    try:
        df = pd.read_excel(file_path, sheet_name='QamqorData', engine='openpyxl', dtype={'БИН': str})
        
        date_columns = ['Дата регистрации', 'Дата акта', 'Дата начала', 'Дата окончания',
                       'Дата приостановки', 'Дата возобновления', 'Начало продления', 'Конец продления']
        
        number_columns = ['Сумма налогов', 'Сумма НДС']
        
        for col in df.columns:
            if col == 'БИН':
                df[col] = df[col].apply(lambda x: normalize_value(x, is_bin=True))
            elif col in date_columns:
                df[col] = pd.to_datetime(df[col], format='%d.%m.%Y', errors='coerce')
            elif col in number_columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            else:
                df[col] = df[col].apply(normalize_value)
        
        df['hash'] = df.apply(lambda row: calculate_row_hash(row.to_dict()), axis=1)
        
        logger.info(f"Загружено {len(df)} записей из Excel")
        return df
    except Exception as e:
        logger.error(f"Ошибка чтения файла: {e}")
        return pd.DataFrame()

def fetch_all_data_from_db(conn_qamqor, conn_companies):
    """Получить все данные из БД"""
    logger.info("Получение данных из БД...")
    
    query_tax = """
    SELECT 
        registration_number, reg_date, act_date, start_date, end_date,
        suspend_date, resume_date, prolong_start, prolong_end,
        revenue_name, kpssu_name, check_type, subject_bin, subject_name,
        subject_address, status, audit_theme, audit_theme_1
    FROM qamqor_tax
    ORDER BY start_date DESC NULLS LAST
    """
    
    df_tax = pd.read_sql(query_tax, conn_qamqor)
    df_tax['check_source'] = 'Налоговая'
    logger.info(f"Налоговых записей: {len(df_tax)}")
    
    query_customs = """
    SELECT 
        registration_number, reg_date, act_date, start_date, end_date,
        suspend_date, resume_date, prolong_start, prolong_end,
        revenue_name, kpssu_name, check_type, subject_bin, subject_name,
        subject_address, status
    FROM qamqor_customs
    ORDER BY start_date DESC NULLS LAST
    """
    
    df_customs = pd.read_sql(query_customs, conn_qamqor)
    df_customs['check_source'] = 'Таможенная'
    for col in ['audit_theme', 'audit_theme_1']:
        df_customs[col] = ''
    logger.info(f"Таможенных записей: {len(df_customs)}")
    
    df_all = pd.concat([df_tax, df_customs], ignore_index=True)
    
    logger.info("Обогащение данными компаний...")
    unique_bins = df_all['subject_bin'].dropna().unique()
    unique_bins_normalized = [normalize_bin(b) for b in unique_bins if normalize_bin(b)]
    companies_data = get_companies_data_batch(unique_bins_normalized, conn_companies)
    
    all_data = []
    for idx, row in df_all.iterrows():
        bin_number = normalize_bin(row['subject_bin'])
        comp_data = companies_data.get(bin_number, {
            'ceo_name': '',
            'krp_name': '',
            'kfc_name': '',
            'kse_name': '',
            'oked_name': '',
            'is_nds': '',
            'tax_history': '',
            'tax_total': None,
            'nds_history': '',
            'nds_total': None,
        })
        
        data_row = {
            'Номер': normalize_value(row['registration_number']),
            'Статус': 'current',
            'Тип проверки': normalize_value(row['check_source']),
            'Дата регистрации': row['reg_date'],
            'Дата акта': row['act_date'],
            'Дата начала': row['start_date'],
            'Дата окончания': row['end_date'],
            'Дата приостановки': row['suspend_date'],
            'Дата возобновления': row['resume_date'],
            'Начало продления': row['prolong_start'],
            'Конец продления': row['prolong_end'],
            'ОГД': shorten_revenue_name(row['revenue_name']),
            'КПССУ': shorten_kpssu_name(row['kpssu_name']),
            'Вид проверки': normalize_value(row['check_type']),
            'БИН': bin_number,
            'Название субъекта': shorten_company_form(row['subject_name']),
            'Адрес субъекта': normalize_value(row['subject_address']),
            'Статус проверки': normalize_value(row['status']),
            'Тема аудита': normalize_value(row.get('audit_theme', '')),
            'Тема аудита 1': normalize_value(row.get('audit_theme_1', '')),
            'Руководитель': comp_data['ceo_name'],
            'КРП': comp_data['krp_name'],
            'КФК': comp_data['kfc_name'],
            'КСЭ': comp_data['kse_name'],
            'ОКЭД': comp_data['oked_name'],
            'Плательщик НДС': comp_data['is_nds'],
            'История налогов': comp_data['tax_history'],
            'Сумма налогов': comp_data['tax_total'],
            'История НДС': comp_data['nds_history'],
            'Сумма НДС': comp_data['nds_total'],
        }
        
        all_data.append(data_row)
    
    df_result = pd.DataFrame(all_data)
    df_result['hash'] = df_result.apply(lambda row: calculate_row_hash(row.to_dict()), axis=1)
    
    logger.info(f"Обработано записей: {len(all_data)}")
    return df_result

# ==================== ОБНОВЛЕНИЕ ДАННЫХ С OPENPYXL ====================

def update_excel_with_openpyxl(file_path, db_df, existing_df):
    """Обновить Excel инкрементально с использованием openpyxl"""
    logger.info("Обновление данных в Excel...")
    
    if not os.path.exists(file_path):
        logger.info("Создание нового файла...")
        db_df['Статус'] = 'new'
        create_new_file(file_path, db_df)
        return len(db_df), 0, 0
    
    wb = load_workbook(file_path)
    ws = wb['QamqorData']
    
    date_columns = ['Дата регистрации', 'Дата акта', 'Дата начала', 'Дата окончания',
                   'Дата приостановки', 'Дата возобновления', 'Начало продления', 'Конец продления']
    
    def write_row_to_excel(ws, row_idx, row_dict):
        """Записать строку в Excel (БЕЗ hash)"""
        for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
            value = row_dict[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_name in date_columns:
                if pd.notna(value):
                    if isinstance(value, pd.Timestamp):
                        cell.value = value.to_pydatetime()
                    elif isinstance(value, (date, datetime)):
                        cell.value = value
                    else:
                        cell.value = ''
                    
                    if cell.value:
                        cell.number_format = 'DD.MM.YYYY'
                else:
                    cell.value = ''
            elif col_name == 'БИН':
                cell.value = value
                cell.number_format = '@'
            elif col_name in ['Сумма налогов', 'Сумма НДС']:
                if pd.notna(value) and value != '':
                    cell.value = float(value) if value else None
                    if cell.value:
                        cell.number_format = '#,##0.00'
                else:
                    cell.value = None
            else:
                cell.value = value
    
    existing_index = {}
    for _, row in existing_df.iterrows():
        record_number = row['Номер']
        record_hash = row['hash']
        existing_index[record_number] = {
            'hash': record_hash
        }
    
    excel_row_index = {}
    for row_idx in range(2, ws.max_row + 1):
        record_number = ws.cell(row=row_idx, column=1).value
        if record_number:
            excel_row_index[record_number] = row_idx
    
    new_count = 0
    updated_count = 0
    unchanged_count = 0
    
    for _, db_row in db_df.iterrows():
        db_row_dict = db_row.to_dict()
        record_id = db_row_dict['Номер']
        new_hash = db_row_dict['hash']
        
        if record_id not in existing_index:
            new_count += 1
            db_row_dict['Статус'] = 'new'
            new_row_idx = ws.max_row + 1
            write_row_to_excel(ws, new_row_idx, db_row_dict)
        else:
            old_hash = existing_index[record_id]['hash']
            
            if new_hash != old_hash:
                updated_count += 1
                db_row_dict['Статус'] = 'updated'
                row_idx = excel_row_index[record_id]
                write_row_to_excel(ws, row_idx, db_row_dict)
            else:
                unchanged_count += 1
    
    table = ws.tables['QamqorData']
    new_max_row = ws.max_row
    table.ref = f"A1:{get_column_letter(len(ALL_COLUMNS))}{new_max_row}"
    
    wb.save(file_path)
    wb.close()
    
    logger.info(f"Новых: {new_count}")
    logger.info(f"Обновлённых: {updated_count}")
    logger.info(f"Без изменений: {unchanged_count}")
    
    return new_count, updated_count, unchanged_count

def create_new_file(file_path, df):
    """Создать новый Excel файл с умной таблицей (БЕЗ колонки hash)"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'QamqorData'
    
    date_columns = ['Дата регистрации', 'Дата акта', 'Дата начала', 'Дата окончания',
                   'Дата приостановки', 'Дата возобновления', 'Начало продления', 'Конец продления']
    
    for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
            value = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_name in date_columns:
                if pd.notna(value):
                    if isinstance(value, pd.Timestamp):
                        cell.value = value.to_pydatetime()
                    elif isinstance(value, (date, datetime)):
                        cell.value = value
                    else:
                        cell.value = ''
                    
                    if cell.value:
                        cell.number_format = 'DD.MM.YYYY'
                else:
                    cell.value = ''
            elif col_name == 'БИН':
                cell.value = value
                cell.number_format = '@'
            elif col_name in ['Сумма налогов', 'Сумма НДС']:
                if pd.notna(value) and value != '':
                    cell.value = float(value) if value else None
                    if cell.value:
                        cell.number_format = '#,##0.00'
                else:
                    cell.value = None
            else:
                cell.value = value
    
    max_row = len(df) + 1
    tab = Table(displayName='QamqorData', ref=f"A1:{get_column_letter(len(ALL_COLUMNS))}{max_row}")
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    ws.freeze_panes = 'A2'
    
    wb.save(file_path)
    wb.close()
    
    logger.info(f"Создан новый файл с {len(df)} записями")

# ==================== MAIN ====================

def main():
    """Основная функция синхронизации"""
    logger.info("=" * 60)
    logger.info("СИНХРОНИЗАЦИЯ БД -> EXCEL")
    logger.info("=" * 60)
    
    conn_qamqor = None
    conn_companies = None
    
    try:
        conn_qamqor = psycopg2.connect(**DB_QAMQOR)
        conn_companies = psycopg2.connect(**DB_COMPANIES)
        
        # ШАГ 1: Сбросить все статусы на 'current'
        reset_all_statuses_to_current(OUTPUT_FILE)
        
        # ШАГ 2: Загрузить существующие данные
        existing_df = load_existing_data(OUTPUT_FILE)
        
        # ШАГ 3: Получить данные из БД
        db_df = fetch_all_data_from_db(conn_qamqor, conn_companies)
        
        # ШАГ 4: Обновить Excel инкрементально
        new_count, updated_count, unchanged_count = update_excel_with_openpyxl(
            OUTPUT_FILE, db_df, existing_df
        )
        
        # Статистика
        total_records = new_count + updated_count + unchanged_count
        
        logger.info("=" * 60)
        logger.info("ИТОГО")
        logger.info("=" * 60)
        logger.info(f"Всего записей: {total_records}")
        logger.info(f"Новых (new): {new_count}")
        logger.info(f"Обновлённых (updated): {updated_count}")
        logger.info(f"Без изменений (current): {unchanged_count}")
        
        type_counts = db_df['Тип проверки'].value_counts()
        logger.info("По типам проверок:")
        for check_type, count in type_counts.items():
            logger.info(f"  {check_type}: {count}")
        
        db_df['Сумма налогов'] = pd.to_numeric(db_df['Сумма налогов'], errors='coerce')
        db_df['Сумма НДС'] = pd.to_numeric(db_df['Сумма НДС'], errors='coerce')
        
        total_tax = db_df['Сумма налогов'].sum()
        total_nds = db_df['Сумма НДС'].sum()
        companies_with_tax = db_df['Сумма налогов'].notna().sum()
        companies_with_nds = db_df['Сумма НДС'].notna().sum()
        
        logger.info("Финансовая статистика:")
        logger.info(f"  Компаний с данными по налогам: {companies_with_tax}")
        logger.info(f"  Общая сумма налогов: {format_number_with_spaces(total_tax)} тенге")
        logger.info(f"  Компаний с данными по НДС: {companies_with_nds}")
        logger.info(f"  Общая сумма НДС: {format_number_with_spaces(total_nds)} тенге")
        
        logger.info("Синхронизация завершена")
        logger.info(f"Файл: {OUTPUT_FILE}")
        
        logger.info("Ожидание синхронизации OneDrive (5 сек)...")
        import time
        time.sleep(5)
        logger.info("Готово")
        
    except Exception as e:
        logger.error(f"Ошибка: {e}", exc_info=True)
        sys.exit(1)
    
    finally:
        if conn_qamqor:
            conn_qamqor.close()
        if conn_companies:
            conn_companies.close()
        logger.info("Подключения к БД закрыты")
        
        import gc
        import time
        
        gc.collect()
        logger.info("Память Python очищена")
        
        logger.info("Ожидание стабилизации файла (10 сек)...")
        time.sleep(10)
        
        logger.info("Файл освобождён и готов к синхронизации OneDrive")
        logger.info("Power Automate обработает изменения в течение 3-5 минут")
        logger.info("=" * 60)

if __name__ == "__main__":
    main()