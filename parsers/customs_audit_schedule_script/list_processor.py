"""
Парсер налогов (КГД) + КФС/ОКЭД (Параграф) по БИН из Excel.

Колонки входного файла: №, ДГД, Компания, БИН (D), Адрес
Добавляет: Tax_sum, Tax_per_year, kfc_code, kfc_name, oked_code, oked_name, Status

- Параллельные запросы KGD + PRG (одновременно)
- Сохраняет каждые SAVE_EVERY БИН
- При повторном запуске продолжает с необработанных (по колонке Status)
- Жесткие таймауты, retry на 5xx
- Игнор ложных SIGINT (от QuickEdit/PowerShell)
"""

import os
import re
import sys
import shutil
import signal
import requests
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutTimeout
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


# Игнорируем все SIGINT (включая ложные от QuickEdit/PowerShell)
signal.signal(signal.SIGINT, signal.SIG_IGN)
try:
    signal.signal(signal.SIGBREAK, signal.SIG_IGN)
except AttributeError:
    pass


# ════════════════════════════════════════════════════════════════
# ⚙️ НАСТРОЙКИ
# ════════════════════════════════════════════════════════════════

INPUT_FILE = "2nd_half_2026.xlsx"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, INPUT_FILE)

BIN_COLUMN = 4          # D
HEADER_ROW = 1
SHEET_NAME = None
LIMIT = None
SAVE_EVERY = 25
CREATE_COPY = True

# Таймауты (connect, read) в секундах
KGD_CONNECT_TIMEOUT = 5
KGD_READ_TIMEOUT = 15
PRG_CONNECT_TIMEOUT = 5
PRG_READ_TIMEOUT = 20
MAX_RETRIES = 1                    # дублировать retry адаптера не нужно
PER_BIN_TIMEOUT = 35               # жесткий таймаут на оба запроса параллельно

NEW_COLUMNS = [
    "Tax_sum",
    "Tax_per_year",
    "kfc_code",
    "kfc_name",
    "oked_code",
    "oked_name",
    "Status",
]


# ════════════════════════════════════════════════════════════════
# 🌐 API
# ════════════════════════════════════════════════════════════════

KGD_TOKEN = "4dfbf650-ef3b-412a-8f4c-5dfc33bcea89"
KGD_URL = "https://portal.kgd.gov.kz/services/isnaportalsync/public/taxpayer-data"
KGD_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36 Edg/144.0.0.0",
    "Accept": "application/json",
    "X-Portal-Token": KGD_TOKEN,
}

PRG_URL = "https://apiba.prgapp.kz/CompanyFullInfo"
PRG_HEADERS = {
    "accept": "application/json",
    "content-type": "application/json",
    "origin": "https://ba.prg.kz",
    "referer": "https://ba.prg.kz/",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
}

BIN_PATTERN = re.compile(r"^\d{12}$")


# ════════════════════════════════════════════════════════════════
# 🌐 СЕССИИ — отдельные для каждого хоста (для thread-safety и keep-alive)
# ════════════════════════════════════════════════════════════════

def make_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=1,
        connect=1,
        read=1,
        backoff_factor=0.3,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET"],
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=5, pool_maxsize=5)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


# ════════════════════════════════════════════════════════════════
# 📡 КГД — налоги
# ════════════════════════════════════════════════════════════════

def fetch_kgd(bin_value: str, session: requests.Session) -> dict:
    out = {"tax_sum": None, "tax_per_year": None, "ok": False, "error": None}
    params = {"taxpayerCode": bin_value, "taxpayerType": "UL"}

    last_err = None
    for _ in range(MAX_RETRIES):
        try:
            resp = session.get(
                KGD_URL,
                headers=KGD_HEADERS,
                params=params,
                timeout=(KGD_CONNECT_TIMEOUT, KGD_READ_TIMEOUT),
            )
            if resp.status_code == 200:
                data = resp.json()
                answer = data.get("answerDto") or {}
                types = ["firstType", "secondType", "thirdType", "ndsType"]

                year_totals = {}
                for t in types:
                    block = answer.get(t) or {}
                    for year, val in block.items():
                        try:
                            year_totals[year] = year_totals.get(year, 0) + float(val or 0)
                        except (TypeError, ValueError):
                            pass

                if year_totals:
                    total = sum(year_totals.values())
                    sorted_years = sorted(year_totals.keys())
                    per_year_str = "\n".join(
                        f"{y}: {year_totals[y]:,.2f}".replace(",", " ")
                        for y in sorted_years
                    )
                    out["tax_sum"] = round(total, 2)
                    out["tax_per_year"] = per_year_str

                out["ok"] = True
                return out
            elif resp.status_code in (500, 502, 503, 504):
                last_err = f"HTTP {resp.status_code}"
                continue
            else:
                out["error"] = f"HTTP {resp.status_code}"
                return out
        except requests.Timeout:
            last_err = "timeout"
            continue
        except requests.RequestException as e:
            last_err = f"net: {type(e).__name__}"
            continue
        except Exception as e:
            last_err = f"err: {type(e).__name__}"
            continue

    out["error"] = last_err or "unknown"
    return out


# ════════════════════════════════════════════════════════════════
# 📡 Параграф — КФС/ОКЭД
# ════════════════════════════════════════════════════════════════

def fetch_prg(bin_value: str, session: requests.Session) -> dict:
    out = {
        "kfc_code": None, "kfc_name": None,
        "oked_code": None, "oked_name": None,
        "ok": False, "error": None, "not_found": False,
    }
    params = {"id": bin_value, "lang": "ru"}

    last_err = None
    for _ in range(MAX_RETRIES):
        try:
            resp = session.get(
                PRG_URL,
                headers=PRG_HEADERS,
                params=params,
                timeout=(PRG_CONNECT_TIMEOUT, PRG_READ_TIMEOUT),
            )
            if resp.status_code == 200:
                data = resp.json()
                basic = data.get("basicInfo") or {}

                if basic.get("isDeleted"):
                    out["not_found"] = True
                    out["ok"] = True
                    out["error"] = "deleted"
                    return out

                kfc = basic.get("kfc")
                if isinstance(kfc, dict):
                    lvl = kfc.get("value")
                    if isinstance(lvl, dict):
                        code = lvl.get("value")
                        desc = lvl.get("description")
                        if code is not None:
                            try:
                                out["kfc_code"] = int(code)
                            except (TypeError, ValueError):
                                out["kfc_code"] = code
                        if desc:
                            out["kfc_name"] = str(desc)

                oked = basic.get("primaryOKED")
                if isinstance(oked, dict):
                    val = oked.get("value")
                    if isinstance(val, dict):
                        val = val.get("value")
                    if isinstance(val, str):
                        parts = val.split(" ", 1)
                        out["oked_code"] = parts[0]
                        out["oked_name"] = parts[1] if len(parts) > 1 else None

                out["ok"] = True
                return out

            elif resp.status_code == 404:
                out["not_found"] = True
                out["ok"] = True
                return out
            elif resp.status_code in (500, 502, 503, 504):
                last_err = f"HTTP {resp.status_code}"
                continue
            else:
                out["error"] = f"HTTP {resp.status_code}"
                return out

        except requests.Timeout:
            last_err = "timeout"
            continue
        except requests.RequestException as e:
            last_err = f"net: {type(e).__name__}"
            continue
        except Exception as e:
            last_err = f"err: {type(e).__name__}"
            continue

    out["error"] = last_err or "unknown"
    return out


# ════════════════════════════════════════════════════════════════
# 📋 EXCEL
# ════════════════════════════════════════════════════════════════

def ensure_headers(filepath, sheet_name=None) -> dict:
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    existing = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v is not None:
            existing[str(v).strip()] = c

    next_col = max_col + 1
    col_idx = {}
    for name in NEW_COLUMNS:
        if name in existing:
            col_idx[name] = existing[name]
        else:
            cell = ws.cell(row=HEADER_ROW, column=next_col, value=name)
            cell.font = Font(bold=True)
            col_idx[name] = next_col
            next_col += 1

    widths = {
        "Tax_sum": 18,
        "Tax_per_year": 28,
        "kfc_code": 10,
        "kfc_name": 50,
        "oked_code": 12,
        "oked_name": 50,
        "Status": 14,
    }
    for name, w in widths.items():
        letter = ws.cell(row=1, column=col_idx[name]).column_letter
        ws.column_dimensions[letter].width = w

    wb.save(filepath)
    wb.close()
    return col_idx


def read_tasks(filepath, col_idx, sheet_name=None) -> list:
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    bin_col = BIN_COLUMN
    status_col = col_idx["Status"]
    max_col_needed = max(bin_col, status_col)

    tasks = []
    skipped_done = 0
    invalid = 0

    for row_cells in ws.iter_rows(
        min_row=HEADER_ROW + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=max_col_needed,
        values_only=False,
    ):
        bin_cell = row_cells[bin_col - 1].value
        if bin_cell is None or str(bin_cell).strip() == "":
            continue

        bin_str = str(bin_cell).strip()
        if "." in bin_str:
            try:
                bin_str = str(int(float(bin_str)))
            except (ValueError, OverflowError):
                pass
        bin_str = bin_str.zfill(12)

        if not BIN_PATTERN.match(bin_str):
            invalid += 1
            continue

        status = row_cells[status_col - 1].value
        if status and str(status).strip().upper() in ("OK", "PARTIAL", "NOT_FOUND"):
            skipped_done += 1
            continue

        row_num = row_cells[0].row
        tasks.append((row_num, bin_str))

    wb.close()
    print(f"📋 Всего к обработке: {len(tasks)}")
    if skipped_done:
        print(f"⏭️  Пропущено (уже обработано): {skipped_done}")
    if invalid:
        print(f"⚠️  Невалидных БИН: {invalid}")
    return tasks


def write_batch(filepath, results, col_idx, sheet_name=None):
    if not results:
        return
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    for r in results:
        row = r["row"]
        ws.cell(row=row, column=col_idx["Tax_sum"], value=r.get("tax_sum"))

        cell = ws.cell(row=row, column=col_idx["Tax_per_year"], value=r.get("tax_per_year"))
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.cell(row=row, column=col_idx["kfc_code"], value=r.get("kfc_code"))
        ws.cell(row=row, column=col_idx["kfc_name"], value=r.get("kfc_name"))
        ws.cell(row=row, column=col_idx["oked_code"], value=r.get("oked_code"))
        ws.cell(row=row, column=col_idx["oked_name"], value=r.get("oked_name"))

        st_cell = ws.cell(row=row, column=col_idx["Status"], value=r.get("status"))
        if r.get("status") == "OK":
            st_cell.font = Font(color="006100")
        elif r.get("status") == "PARTIAL":
            st_cell.font = Font(color="9C5700")
        elif r.get("status") in ("ERROR", "NOT_FOUND"):
            st_cell.font = Font(color="C00000")

    wb.save(filepath)
    wb.close()


# ════════════════════════════════════════════════════════════════
# 🚀 MAIN
# ════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  ПАРСЕР: КГД (налоги) + Параграф (КФС/ОКЭД)")
    print("=" * 60)

    if not os.path.exists(INPUT_FILE):
        print(f"❌ Файл не найден: {INPUT_FILE}")
        sys.exit(1)

    if CREATE_COPY:
        p = Path(INPUT_FILE)
        output_file = str(p.parent / f"{p.stem}_result{p.suffix}")
        if not os.path.exists(output_file):
            shutil.copy2(INPUT_FILE, output_file)
            print(f"📄 Создана копия: {output_file}")
        else:
            print(f"📄 Используется существующий файл: {output_file}")
    else:
        output_file = INPUT_FILE

    col_idx = ensure_headers(output_file, SHEET_NAME)
    tasks = read_tasks(output_file, col_idx, SHEET_NAME)

    if not tasks:
        print("\n🎉 Все БИН уже обработаны.")
        return

    if LIMIT:
        tasks = tasks[:LIMIT]
        print(f"🔢 Лимит: {LIMIT}")

    # Отдельные сессии под каждый хост (изолированные пулы соединений)
    kgd_session = make_session()
    prg_session = make_session()

    # Один пул потоков на весь скрипт
    executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="fetcher")

    buffer = []
    stats = {"ok": 0, "partial": 0, "error": 0, "not_found": 0, "total": len(tasks)}
    saved_total = 0

    try:
        for i, (row, bin_value) in enumerate(tasks, 1):
            # Параллельный запуск KGD + PRG
            f_kgd = executor.submit(fetch_kgd, bin_value, kgd_session)
            f_prg = executor.submit(fetch_prg, bin_value, prg_session)

            try:
                kgd = f_kgd.result(timeout=PER_BIN_TIMEOUT)
            except FutTimeout:
                kgd = {"tax_sum": None, "tax_per_year": None, "ok": False, "error": "future_timeout"}
            except Exception as e:
                kgd = {"tax_sum": None, "tax_per_year": None, "ok": False, "error": f"exc: {type(e).__name__}"}

            try:
                prg = f_prg.result(timeout=PER_BIN_TIMEOUT)
            except FutTimeout:
                prg = {"kfc_code": None, "kfc_name": None, "oked_code": None, "oked_name": None,
                       "ok": False, "error": "future_timeout", "not_found": False}
            except Exception as e:
                prg = {"kfc_code": None, "kfc_name": None, "oked_code": None, "oked_name": None,
                       "ok": False, "error": f"exc: {type(e).__name__}", "not_found": False}

            # Определяем статус
            if kgd["ok"] and prg["ok"] and not prg.get("not_found"):
                status = "OK"
                stats["ok"] += 1
            elif prg.get("not_found") and kgd["ok"] and kgd["tax_sum"] is None:
                status = "NOT_FOUND"
                stats["not_found"] += 1
            elif kgd["ok"] or prg["ok"]:
                status = "PARTIAL"
                stats["partial"] += 1
            else:
                status = "ERROR"
                stats["error"] += 1

            print(
                f"[{i}/{len(tasks)}] {bin_value} | "
                f"KGD: ok={kgd['ok']} sum={kgd['tax_sum']} err={kgd['error']} | "
                f"PRG: ok={prg['ok']} kfc={prg['kfc_code']} oked={prg['oked_code']} err={prg['error']} "
                f"→ {status}",
                flush=True
            )

            buffer.append({
                "row": row,
                "tax_sum": kgd.get("tax_sum"),
                "tax_per_year": kgd.get("tax_per_year"),
                "kfc_code": prg.get("kfc_code"),
                "kfc_name": prg.get("kfc_name"),
                "oked_code": prg.get("oked_code"),
                "oked_name": prg.get("oked_name"),
                "status": status,
            })

            if len(buffer) >= SAVE_EVERY:
                try:
                    write_batch(output_file, buffer, col_idx, SHEET_NAME)
                    saved_total += len(buffer)
                    buffer.clear()
                    print(f"  💾 Автосохранение ({saved_total}/{len(tasks)})", flush=True)
                except Exception as e:
                    print(f"  ⚠️ Ошибка сохранения: {e}", flush=True)

    except Exception as e:
        print(f"\n\n❌ Критическая ошибка: {type(e).__name__}: {e}")
        raise
    finally:
        # Финальное сохранение
        if buffer:
            try:
                write_batch(output_file, buffer, col_idx, SHEET_NAME)
                saved_total += len(buffer)
                buffer.clear()
            except Exception as e:
                print(f"⚠️ Ошибка финального сохранения: {e}")

        # Завершаем пул
        executor.shutdown(wait=True)
        kgd_session.close()
        prg_session.close()

    print("\n" + "=" * 60)
    print("📊 ИТОГИ")
    print("=" * 60)
    print(f"  Всего:        {stats['total']}")
    print(f"  ✅ OK:         {stats['ok']}")
    print(f"  🟡 PARTIAL:    {stats['partial']}")
    print(f"  ❌ NOT_FOUND:  {stats['not_found']}")
    print(f"  ⚠️  ERROR:      {stats['error']}")
    print(f"  💾 Сохранено:  {saved_total}")
    print(f"\n📁 Результат: {output_file}")


if __name__ == "__main__":
    main()