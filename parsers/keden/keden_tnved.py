#!/usr/bin/env python3
"""
Скрипт для извлечения RESTRICTION-документов из keden_full_response.json
и формирования краткого Excel-отчёта на двух листах (IN / OUT).

Фильтры (идентичны парсеру test.py):
  - Только docType == "RESTRICTION"
  - Исключаются документы, содержащие в name: ветеринар, животн, лекарств, медицинс

Колонки:
  A — Код ТН ВЭД ЕАЭС
  B — Наименование ограничения (documents.name)
  C — Примечания (documents.comment + documents.comments[])
  D — Ссылка (resolution.url)
"""

import json
import os
import sys
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side


# Фильтр исключений — идентичен парсеру
EXCLUDE_KEYWORDS = ["ветеринар", "животн", "лекарств", "медицинс"]

# Имена файлов
JSON_FILENAME = "keden_full_response.json"
EXCEL_FILENAME = "keden_restrictions_short.xlsx"


def should_exclude_document(doc_name: str) -> bool:
    """Проверяет, нужно ли исключить документ по ключевым словам."""
    if not doc_name:
        return False
    name_lower = doc_name.lower()
    for keyword in EXCLUDE_KEYWORDS:
        if keyword.lower() in name_lower:
            return True
    return False


def format_comments(doc: dict) -> str:
    """
    Собирает comment и comments[] документа в единый текст.

    Формат:
        * comment (если есть)
        - comments[0]
        - comments[1]
        ...
    """
    lines = []

    # comment — одиночный
    comment = doc.get("comment")
    if comment is not None:
        comment = str(comment).strip()
        if comment:
            lines.append(f"* {comment}")

    # comments[] — массив
    comments = doc.get("comments") or []
    for item in comments:
        if item is None:
            continue
        item = str(item).strip()
        if item:
            lines.append(f"- {item}")

    return "\n".join(lines)


def extract_restrictions_by_direction(data: dict, direction: str) -> list:
    """
    Извлекает четвёрки (код, documents.name, примечания, url) для указанного направления.
    Каждый documents.name — отдельная запись.

    Возвращает:
        [
            ("3919900000", "Входит в перечень...", "* примечание\n- комментарий", "https://..."),
            ...
        ]
    """
    rows = []

    for code in sorted(data.keys()):
        code_data = data[code]
        if code_data is None or not isinstance(code_data, dict):
            continue

        direction_data = code_data.get(direction)
        if not direction_data or not isinstance(direction_data, dict):
            continue

        non_tariff_resolutions = direction_data.get("nonTariffResolutions") or []

        seen_names = set()

        for resolution in non_tariff_resolutions:
            documents = resolution.get("documents") or []
            res_url = (resolution.get("url") or "").strip()

            for doc in documents:
                doc_type = (doc.get("docType") or "").strip().upper()
                if doc_type != "RESTRICTION":
                    continue

                doc_name = (doc.get("name") or "").strip()
                if not doc_name:
                    continue

                if should_exclude_document(doc_name):
                    continue

                # Убираем дубли внутри одного кода
                if doc_name in seen_names:
                    continue
                seen_names.add(doc_name)

                # Собираем примечания
                comments_text = format_comments(doc)

                rows.append((code, doc_name, comments_text, res_url))

    return rows


def create_sheet(wb: openpyxl.Workbook, sheet_name: str, rows: list) -> int:
    """Создаёт лист Excel и заполняет его данными."""
    ws = wb.create_sheet(title=sheet_name)

    # Стили
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    code_alignment = Alignment(vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Заголовки
    headers = [
        "Код ТН ВЭД ЕАЭС",
        "Наименование ограничения",
        "Примечания (comment / comments)",
        "Ссылка (url)",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Данные — каждый documents.name в отдельной строке
    for i, (code, doc_name, comments_text, url) in enumerate(rows):
        row_num = i + 2

        cell_code = ws.cell(row=row_num, column=1, value=code)
        cell_code.alignment = code_alignment
        cell_code.border = thin_border

        cell_doc = ws.cell(row=row_num, column=2, value=doc_name)
        cell_doc.alignment = cell_alignment
        cell_doc.border = thin_border

        cell_comments = ws.cell(row=row_num, column=3, value=comments_text if comments_text else "")
        cell_comments.alignment = cell_alignment
        cell_comments.border = thin_border

        cell_url = ws.cell(row=row_num, column=4, value=url if url else "")
        cell_url.alignment = cell_alignment
        cell_url.border = thin_border

    # Ширина колонок
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 50
    ws.freeze_panes = "A2"

    return len(rows)


def save_excel(data: dict, filepath: str):
    """Формирует Excel с двумя листами: IN (импорт) и OUT (экспорт)."""

    in_rows = extract_restrictions_by_direction(data, "IN")
    out_rows = extract_restrictions_by_direction(data, "OUT")

    if not in_rows and not out_rows:
        print("[EXCEL] ⚠ Нет RESTRICTION-документов — файл не создан")
        return

    wb = openpyxl.Workbook()
    # Удаляем дефолтный лист
    wb.remove(wb.active)

    # Лист 1 — Импорт (IN)
    in_count = create_sheet(wb, "Импорт (IN)", in_rows)
    print(f"  Лист 'Импорт (IN)':   {in_count} строк")

    # Лист 2 — Экспорт (OUT)
    out_count = create_sheet(wb, "Экспорт (OUT)", out_rows)
    print(f"  Лист 'Экспорт (OUT)': {out_count} строк")

    wb.save(filepath)
    print(f"[EXCEL] ✅ Сохранено: {filepath}")


def find_json_file() -> str:
    """Ищет JSON-файл рядом со скриптом или в рабочей директории."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, JSON_FILENAME)

    if os.path.isfile(json_path):
        return json_path

    cwd_path = os.path.join(os.getcwd(), JSON_FILENAME)
    if os.path.isfile(cwd_path):
        print(f"[INFO] Файл найден в рабочей директории: {cwd_path}")
        return cwd_path

    print(f"[ОШИБКА] Файл '{JSON_FILENAME}' не найден!")
    print(f"  Искали в:")
    print(f"    1. {json_path}")
    print(f"    2. {cwd_path}")
    print(f"\n  Убедитесь, что файл '{JSON_FILENAME}' лежит рядом со скриптом.")
    sys.exit(1)


def load_json(filepath: str) -> dict:
    """Загружает и парсит JSON-файл с обработкой ошибок."""
    print(f"[JSON] Загрузка: {filepath}")

    try:
        size = os.path.getsize(filepath)
        if size == 0:
            print(f"[ОШИБКА] Файл пуст (0 байт): {filepath}")
            sys.exit(1)
        print(f"[JSON] Размер файла: {size:,} байт")
    except OSError as e:
        print(f"[ОШИБКА] Не удалось прочитать файл: {e}")
        sys.exit(1)

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"[ОШИБКА] Невалидный JSON: {e}")
        sys.exit(1)
    except UnicodeDecodeError as e:
        print(f"[ОШИБКА] Проблема с кодировкой файла: {e}")
        sys.exit(1)
    except PermissionError:
        print(f"[ОШИБКА] Нет прав на чтение файла: {filepath}")
        sys.exit(1)
    except Exception as e:
        print(f"[ОШИБКА] Неожиданная ошибка: {type(e).__name__}: {e}")
        sys.exit(1)

    if not isinstance(data, dict):
        print(f"[ОШИБКА] Ожидался JSON-объект (dict), получен: {type(data).__name__}")
        sys.exit(1)

    print(f"[JSON] ✅ Загружено кодов: {len(data)}")
    return data


def main():
    print("=" * 60)
    print("Извлечение RESTRICTION из keden_full_response.json")
    print(f"Фильтр исключений: {', '.join(EXCLUDE_KEYWORDS)}")
    print("=" * 60)

    # 1. Найти JSON
    json_path = find_json_file()

    # 2. Загрузить JSON
    data = load_json(json_path)

    # 3. Статистика
    in_rows = extract_restrictions_by_direction(data, "IN")
    out_rows = extract_restrictions_by_direction(data, "OUT")

    in_codes = len(set(code for code, _, _, _ in in_rows))
    out_codes = len(set(code for code, _, _, _ in out_rows))

    in_with_comments = sum(1 for _, _, c, _ in in_rows if c)
    out_with_comments = sum(1 for _, _, c, _ in out_rows if c)

    print(f"\n[ИТОГО]")
    print(f"  Всего кодов в JSON:              {len(data)}")
    print(f"  Импорт — кодов с RESTRICTION:    {in_codes} ({len(in_rows)} документов, {in_with_comments} с примечаниями)")
    print(f"  Экспорт — кодов с RESTRICTION:   {out_codes} ({len(out_rows)} документов, {out_with_comments} с примечаниями)")

    # 4. Сохранить Excel
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILENAME)

    try:
        print(f"\n[EXCEL] Формирование: {excel_path}")
        save_excel(data, excel_path)
    except PermissionError:
        print(f"[ОШИБКА] Файл '{EXCEL_FILENAME}' открыт в другой программе!")
        print(f"  Закройте его и запустите скрипт заново.")
        sys.exit(1)
    except Exception as e:
        print(f"[ОШИБКА] Ошибка сохранения: {type(e).__name__}: {e}")
        sys.exit(1)

    print(f"\n✅ Готово!")


if __name__ == "__main__":
    main()