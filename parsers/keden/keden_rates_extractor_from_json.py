"""
Скрипт для преобразования JSON-ответа Кеден в Excel-таблицу со ставками.

Извлекает из JSON структуры данные о ставках (rates) по каждому коду товара
и каждой дате, игнорируя блок nonTariffResolutions.
"""

import json
import logging
import re
import sys
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --- Конфигурация ---------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_FILE = SCRIPT_DIR / "keden_full_response.json"
OUTPUT_FILE = SCRIPT_DIR / "keden_rates.xlsx"
SHEET_NAME = "Rates"

COLUMNS = ("Код", "Дата", "Ставка", "Комментарий к ставке", "Дата начала ставки")

# Управляющие символы, запрещённые в ячейках xlsx (кроме \t \n \r).
# Соответствует openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE.
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


# --- Логирование ----------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)


# --- Утилиты --------------------------------------------------------------

def clean_cell(value: Any) -> Any:
    """Удаляет управляющие символы, недопустимые в ячейках xlsx."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_RE.sub("", value)
    return value


# --- Бизнес-логика --------------------------------------------------------

def load_json(path: Path) -> dict[str, Any]:
    """Загружает JSON-файл."""
    logger.info("Чтение файла: %s", path)
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def iter_rate_rows(data: dict[str, Any]) -> Iterator[tuple[str, str, str, str, str]]:
    """
    Обходит структуру JSON и выдаёт строки для Excel.

    Структура JSON:
        { "<код>": { "<дата>": { "IN": { "rates": [ {...}, ... ] } } } }
    """
    for code, dates in data.items():
        for date, payload in dates.items():
            rates = payload.get("IN", {}).get("rates", [])
            for rate in rates:
                yield (
                    clean_cell(code),
                    clean_cell(date),
                    clean_cell(rate.get("rateValue", "")),
                    clean_cell(rate.get("comment", "")),
                    clean_cell(rate.get("beginDate", "")),
                )


# --- Работа с Excel -------------------------------------------------------

def style_header(ws: Worksheet) -> None:
    """Применяет оформление к заголовку таблицы."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill


def autosize_columns(ws: Worksheet, max_width: int = 60) -> None:
    """Автоподбор ширины колонок (с ограничением сверху)."""
    for col_idx, _ in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def write_xlsx(rows: list[tuple], output_path: Path) -> None:
    """Записывает данные в xlsx-файл."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(list(COLUMNS))
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    style_header(ws)
    autosize_columns(ws)

    wb.save(output_path)
    logger.info("Файл сохранён: %s (строк: %d)", output_path, len(rows))


# --- Точка входа ----------------------------------------------------------

def main() -> int:
    if not INPUT_FILE.exists():
        logger.error("Входной файл не найден: %s", INPUT_FILE)
        return 1

    try:
        data = load_json(INPUT_FILE)
    except json.JSONDecodeError as e:
        logger.error("Ошибка парсинга JSON: %s", e)
        return 1

    rows = list(iter_rate_rows(data))
    if not rows:
        logger.warning("В JSON не найдено ни одной ставки.")

    write_xlsx(rows, OUTPUT_FILE)
    return 0


if __name__ == "__main__":
    sys.exit(main())