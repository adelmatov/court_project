#!/usr/bin/env python3
"""
Простой парсер KFC для Excel файла с БИН.

Использование:
    python kfc_parser.py input.xlsx

Входной файл: Excel с БИН в первой колонке
Выходной файл: тот же файл с KFC во второй колонке
"""

import sys
import signal
import time
import random
from pathlib import Path

import requests
import openpyxl
from openpyxl import load_workbook


# ═══════════════════════════════════════════════════════════════
# КОНФИГУРАЦИЯ
# ═══════════════════════════════════════════════════════════════

API_URL = "https://apiba.prgapp.kz/CompanyFullInfo"
API_HEADERS = {
    'accept': 'application/json',
    'origin': 'https://ba.prg.kz',
    'referer': 'https://ba.prg.kz/',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0'
}
MIN_DELAY = 6.0
MAX_DELAY = 10.0
TIMEOUT = 30

# Флаг для корректного завершения
stop_requested = False


def signal_handler(signum, frame):
    """Обработчик сигнала прерывания."""
    global stop_requested
    print("\n⚠️  Получен сигнал остановки. Сохраняю данные...")
    stop_requested = True


# Регистрируем обработчик
signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)


# ═══════════════════════════════════════════════════════════════
# ФУНКЦИИ
# ═══════════════════════════════════════════════════════════════

def get_kfc(bin_value: str) -> str:
    """
    Получить KFC для БИН.
    
    Returns:
        Только описание KFC или "NOT_FOUND" / "ERROR"
    """
    try:
        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))
        
        response = requests.get(
            API_URL,
            params={'id': bin_value, 'lang': 'ru'},
            headers=API_HEADERS,
            timeout=TIMEOUT
        )
        
        if response.status_code == 404:
            return "NOT_FOUND"
        
        if response.status_code != 200:
            return f"ERROR_{response.status_code}"
        
        data = response.json()
        basic = data.get('basicInfo', {})
        
        # Проверка на удалённую компанию
        if basic.get('isDeleted', False):
            return "DELETED"
        
        # Извлечение KFC
        kfc_field = basic.get('kfc')
        if not kfc_field:
            return "NO_KFC"
        
        kfc_value = kfc_field.get('value')
        if not kfc_value or not isinstance(kfc_value, dict):
            return "NO_KFC"
        
        # ✅ Возвращаем ТОЛЬКО описание
        description = kfc_value.get('description', '')
        
        if description:
            return description
        
        return "NO_KFC"
        
    except requests.Timeout:
        return "TIMEOUT"
    except requests.RequestException as e:
        return f"ERROR: {str(e)[:50]}"
    except Exception as e:
        return f"ERROR: {str(e)[:50]}"


def process_excel(filepath: str):
    """Обработать Excel файл."""
    global stop_requested
    
    path = Path(filepath)
    if not path.exists():
        print(f"❌ Файл не найден: {filepath}")
        sys.exit(1)
    
    print(f"📂 Открываю файл: {filepath}")
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Статистика
    total = 0
    processed = 0
    skipped = 0
    
    # Подсчёт строк
    rows_to_process = []
    for row_num in range(1, ws.max_row + 1):
        bin_cell = ws.cell(row=row_num, column=1)
        kfc_cell = ws.cell(row=row_num, column=2)
        
        bin_value = str(bin_cell.value or '').strip()
        kfc_value = str(kfc_cell.value or '').strip()
        
        # Проверяем что БИН валидный (12 цифр)
        if bin_value and bin_value.isdigit() and len(bin_value) == 12:
            total += 1
            
            # Пропускаем если KFC уже есть
            if kfc_value and kfc_value not in ('', 'None', 'ERROR', 'TIMEOUT'):
                skipped += 1
            else:
                rows_to_process.append(row_num)
    
    print(f"📊 Всего БИН: {total}")
    print(f"⏭️  Уже обработано: {skipped}")
    print(f"🔄 Осталось обработать: {len(rows_to_process)}")
    print("-" * 50)
    
    if not rows_to_process:
        print("✅ Все БИН уже обработаны!")
        return
    
    # Обработка
    for i, row_num in enumerate(rows_to_process, 1):
        if stop_requested:
            print(f"\n💾 Сохраняю после {processed} обработанных записей...")
            break
        
        bin_cell = ws.cell(row=row_num, column=1)
        kfc_cell = ws.cell(row=row_num, column=2)
        bin_value = str(bin_cell.value).strip()
        
        print(f"[{i}/{len(rows_to_process)}] БИН: {bin_value} ... ", end='', flush=True)
        
        kfc = get_kfc(bin_value)
        kfc_cell.value = kfc
        processed += 1
        
        print(f"KFC: {kfc}")
        
        # Сохраняем каждые 10 записей
        if processed % 10 == 0:
            wb.save(filepath)
            print(f"   💾 Автосохранение ({processed} записей)")
    
    # Финальное сохранение
    wb.save(filepath)
    print("-" * 50)
    print(f"✅ Готово! Обработано: {processed}")
    print(f"📁 Результат сохранён в: {filepath}")


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    # Файл рядом со скриптом
    script_dir = Path(__file__).parent
    filepath = script_dir / "bin_list_2026_1st_half.xlsx"
    
    try:
        process_excel(str(filepath))
    except KeyboardInterrupt:
        print("\n⚠️  Прервано пользователем")
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        sys.exit(1)